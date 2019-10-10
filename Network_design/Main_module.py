
import openpyxl as excell
import values as val

A,b_ij,K,d = val.generate_val()

file = excell.Workbook()
activeSheet = file.active
    
def Dijksters_Short_Cost(node_one,Row,visited,unvisited,p):
    # the statement first visites the current node i.e. Row node and sets cost to '0'
    activeSheet.cell(row=Row,column=Row-p*21).value = str(Row-p*21-1)+','+str(0)
    # initializing the current cost to '0'
    current_cost = 0
    node = Row-1-p*21
    # visiting the unvisited nodes till all the nodes are visited
    while(unvisited):
        # Now taking the node 'Row' i.e. visiting the 'Row'th node
        # and obtaining the minimum cost node
        
        for row in activeSheet.iter_rows(min_row=Row,max_row=Row,min_col=1,max_col=len(d)):
            #node = unvisited[0]
            cost = float(activeSheet.cell(row=Row,column=node+1).value.split(',')[1])
            for RoW in row:
                if(float(RoW.value.split(',')[1])<cost and (int(RoW.value.split(',')[0]) in unvisited)):
                    cost = float(RoW.value.split(',')[1])
                    node = int(RoW.value.split(',')[0])
                    
        current_cost =cost
        current_node = node
        for j in unvisited:
            get_cost = node_one[current_node,j]
            if(get_cost==0):
                continue
                
            if(float(activeSheet.cell(row=Row,column=j+1).value.split(',')[1])>=current_cost+get_cost):
                activeSheet.cell(row=Row,column=j+1).value = \
                str(int(activeSheet.cell(row=Row,column=j+1).value.split(',')[0]))\
                +','+str(int(get_cost+current_cost))+','+str(int(current_node))
                
                
        visited.append(node)
        unvisited.remove(node)
        
        if(unvisited):
            node = unvisited[0]
        
    file.save('ATNexcell.xlsx')
        

def source_node(graph1,p):
    row = 1
    activeSheet.merge_cells(start_row=row+p*21,start_column=23,end_row=row+p*21+4,end_column = 27)
    activeSheet.cell(row=row+p*21,column=23).value = 'K = '+str(p+3)
    for i in range(len(d)):
        for j in range(len(d)):
            # stores (node,cost)
            activeSheet.cell(row=(i+1+p*21),column=j+1).value = str(j)+',inf'
            
    for source in graph1:
        unvisited = list(range(len(d)))
        visited = []
        Dijksters_Short_Cost(graph1,row+p*21,visited,unvisited,p)
        row+=1
        
        
def network(A):
    for p in range(len(A)):
        source_node(A[p],p)
        
def total_cost(p):
    Z_opt =0
    for i in range(len(d)):
        for j in range(len(d)):
            # sum(d_kl*sum(a_ij))
            if(i==j):
                continue
            
            a_ij = int(activeSheet.cell(row=i+1+p*21,column=j+1).value.split(',')[1])
            Z_opt += b_ij[i][j]*a_ij
    return Z_opt

def network_density(p):
    path = 0
    li = []
    for i in range(len(d)):
        for j in range(len(d)):
            if(i==j):
                continue
                
            c = (int(activeSheet.cell(row=i+1+p*21,column=j+1).value.split(',')[0]),
                           int(activeSheet.cell(row=i+1+p*21,column=j+1).value.split(',')[2]))
            li.append(c)
    links = len(set(li))
    return links/(20*19)

def plot_values():
    T_cost = []
    N_dens = []
    network(A)
    for l in range(len(K)):
        T_cost.append(total_cost(l))
        N_dens.append(network_density(l))
        
    return T_cost,N_dens,K